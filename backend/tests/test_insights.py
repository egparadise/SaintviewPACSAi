"""인사이트 API — 시스템 로그(분류·필터·CSV) · 사용량 통계(JSON·Excel) · DB 구조 · DB 도구 · 관리자 생성 가드."""
from __future__ import annotations

import io
import sys
import uuid
from datetime import datetime, timedelta, timezone

from app.api import insights
from app.main import app

# B1 레인이 main.py 에 등록하기 전까지 테스트에서 직접 마운트(중복 등록 방지 가드)
if not any(getattr(r, "path", "") == "/api/insights/stats" for r in app.router.routes):
    app.include_router(insights.router)


# ────────────────────────────── 헬퍼 ──────────────────────────────
def _mk_hospital(client, auth_headers, code: str) -> int:
    r = client.post("/api/admin/hospitals", headers=auth_headers,
                    json={"code": code, "name": f"병원-{code}"})
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _seed_studies(db, hid: int) -> None:
    """통계용 검사 시드 — CT 확정 1 · CT 미판독 1 · MR 미판독 1 (진료과 2종, 2026-06월)."""
    from app.models import Patient, Study

    p = Patient(patient_key=f"PK-{uuid.uuid4().hex[:10]}")
    db.add(p)
    db.flush()

    def mk(modality, status, dept, date8):
        return Study(patient_id=p.id, study_uid=uuid.uuid4().hex, modality=modality,
                     status=status, department=dept, study_date=date8, hospital_id=hid)

    db.add_all([
        mk("CT", "finalized", "영상의학과", "20260601"),
        mk("CT", "received", "영상의학과", "20260610"),
        mk("MR", "received", "신경외과", "20260620"),
        # 기간 필터 밖(범위 검증용)
        mk("US", "received", "내과", "20250101"),
    ])
    db.commit()


def _login(client, username: str, password: str) -> dict:
    r = client.post("/api/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['token']}"}


# ────────────────────────────── 시스템 로그 ──────────────────────────────
def test_logs_classification_and_filters(client, auth_headers):
    # auth_headers 픽스처의 admin 로그인이 'login' 감사 로그를 남긴다 → network 분류
    r = client.get("/api/insights/logs", headers=auth_headers,
                   params={"type": "network", "q": "login"})
    assert r.status_code == 200, r.text
    items = r.json()["items"]
    assert items, "로그인 이벤트가 network 로그로 분류되어야 한다"
    assert all(it["type"] == "network" for it in items)
    assert any(it["action"] == "login" for it in items)
    # 응답 항목 계약 형태
    for k in ("ts", "type", "actor", "hospital_id", "action", "detail"):
        assert k in items[0]

    # event 타입 — network/dicom 액션은 섞이지 않는다
    r = client.get("/api/insights/logs", headers=auth_headers, params={"type": "event"})
    assert r.status_code == 200
    assert all(it["type"] == "event" for it in r.json()["items"])

    # 미래 날짜 필터 → 빈 결과
    tomorrow = (datetime.now(timezone.utc) + timedelta(days=1)).strftime("%Y-%m-%d")
    r = client.get("/api/insights/logs", headers=auth_headers, params={"date_from": tomorrow})
    assert r.status_code == 200
    assert r.json()["items"] == []

    # 잘못된 type / 날짜 → 400
    assert client.get("/api/insights/logs", headers=auth_headers,
                      params={"type": "nope"}).status_code == 400
    assert client.get("/api/insights/logs", headers=auth_headers,
                      params={"date_from": "어제"}).status_code == 400


def test_logs_hid_filter(client, auth_headers):
    hid = _mk_hospital(client, auth_headers, f"HLOG{uuid.uuid4().hex[:4]}")
    # 병원 소속 계정 생성 → 로그인(해당 병원 귀속 network 로그 발생)
    uname = f"hlog_{uuid.uuid4().hex[:6]}"
    r = client.post("/api/admin/accounts", headers=auth_headers, json={
        "username": uname, "password": "pass12345", "role": "staff", "hospital_id": hid,
    })
    assert r.status_code == 200, r.text
    _login(client, uname, "pass12345")
    r = client.get("/api/insights/logs", headers=auth_headers, params={"hid": hid})
    assert r.status_code == 200
    items = r.json()["items"]
    assert items, "병원 스코프 로그가 있어야 한다"
    assert all(it["hospital_id"] == hid for it in items)
    assert any(it["action"] == "login" and it["actor"] == uname for it in items)


def test_logs_csv_download(client, auth_headers):
    r = client.get("/api/insights/logs.csv", headers=auth_headers, params={"type": "network"})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    assert "attachment" in r.headers.get("content-disposition", "")
    body = r.content.decode("utf-8")
    assert body.startswith("﻿"), "엑셀 한글 안전을 위해 BOM 으로 시작해야 한다"
    header = body.lstrip("﻿").splitlines()[0]
    assert header == "ts,type,actor,hospital_id,action,detail"
    assert "login" in body


def test_logs_admin_only(client, auth_headers):
    uname = f"nadm_{uuid.uuid4().hex[:6]}"
    r = client.post("/api/admin/accounts", headers=auth_headers, json={
        "username": uname, "password": "pass12345", "role": "technologist",
    })
    assert r.status_code == 200, r.text
    hdrs = _login(client, uname, "pass12345")
    assert client.get("/api/insights/logs", headers=hdrs).status_code == 403
    assert client.get("/api/insights/stats", headers=hdrs).status_code == 403
    assert client.get("/api/insights/db-schema", headers=hdrs).status_code == 403
    assert client.post("/api/insights/db-tool-open", headers=hdrs).status_code == 403


# ────────────────────────────── 사용량 통계 ──────────────────────────────
def test_stats_groups(client, auth_headers, db):
    hid = _mk_hospital(client, auth_headers, f"HSTA{uuid.uuid4().hex[:4]}")
    _seed_studies(db, hid)
    base = {"hid": hid, "date_from": "2026-06-01", "date_to": "2026-06-30"}

    # 장비(modality)별
    r = client.get("/api/insights/stats", headers=auth_headers,
                   params={**base, "group": "modality"})
    assert r.status_code == 200, r.text
    rows = {x["key"]: x for x in r.json()["rows"]}
    assert rows["CT"]["studies"] == 2 and rows["CT"]["reports"] == 1 and rows["CT"]["unreported"] == 1
    assert rows["MR"]["studies"] == 1 and rows["MR"]["unreported"] == 1
    assert "US" not in rows  # 기간 밖 검사는 제외

    # 진료과(department)별
    r = client.get("/api/insights/stats", headers=auth_headers,
                   params={**base, "group": "department"})
    rows = {x["key"]: x for x in r.json()["rows"]}
    assert rows["영상의학과"]["studies"] == 2
    assert rows["신경외과"]["studies"] == 1

    # 판독 상태(report_status)
    r = client.get("/api/insights/stats", headers=auth_headers,
                   params={**base, "group": "report_status"})
    rows = {x["key"]: x for x in r.json()["rows"]}
    assert rows["finalized"]["studies"] == 1
    assert rows["unreported"]["studies"] == 2

    # 병원(hospital)별 — hid 스코프에서 해당 병원 한 행
    r = client.get("/api/insights/stats", headers=auth_headers,
                   params={**base, "group": "hospital"})
    rows = r.json()["rows"]
    assert len(rows) == 1 and rows[0]["key"] == str(hid) and rows[0]["studies"] == 3

    # 알 수 없는 group → 400
    assert client.get("/api/insights/stats", headers=auth_headers,
                      params={"group": "nope"}).status_code == 400


def test_stats_xlsx_download(client, auth_headers, db):
    """Excel 내보내기 — content-type · PK 매직 바이트 · 시트명(그룹 라벨) · 데이터 행 검증."""
    hid = _mk_hospital(client, auth_headers, f"HXLS{uuid.uuid4().hex[:4]}")
    _seed_studies(db, hid)
    r = client.get("/api/insights/stats.xlsx", headers=auth_headers, params={
        "group": "modality", "hid": hid, "date_from": "2026-06-01", "date_to": "2026-06-30",
    })
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "attachment" in r.headers.get("content-disposition", "")
    assert r.content[:2] == b"PK", "xlsx(zip) 매직 바이트로 시작해야 한다"

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws.title == "장비별"  # 시트명 = 그룹 한국어 라벨
    assert [c.value for c in ws[1]] == ["구분", "검사 수", "판독", "미판독"]
    assert ws[1][0].font.bold, "헤더는 볼드"
    rows = {row[0].value: (row[1].value, row[2].value, row[3].value)
            for row in ws.iter_rows(min_row=2)}
    assert rows["CT"] == (2, 1, 1)  # 시드: CT 확정1+미판독1
    assert rows["MR"] == (1, 0, 1)
    assert rows["합계"] == (3, 1, 2)  # 합계 행

    # 알 수 없는 group → 400 (JSON 경로와 동일 검증 공유)
    assert client.get("/api/insights/stats.xlsx", headers=auth_headers,
                      params={"group": "nope"}).status_code == 400


def test_stats_xlsx_admin_only(client, auth_headers):
    uname = f"nadx_{uuid.uuid4().hex[:6]}"
    r = client.post("/api/admin/accounts", headers=auth_headers, json={
        "username": uname, "password": "pass12345", "role": "technologist",
    })
    assert r.status_code == 200, r.text
    hdrs = _login(client, uname, "pass12345")
    assert client.get("/api/insights/stats.xlsx", headers=hdrs).status_code == 403


# ────────────────────────────── DB 구조 / DB 도구 ──────────────────────────────
def test_db_schema_readonly(client, auth_headers):
    r = client.get("/api/insights/db-schema", headers=auth_headers)
    assert r.status_code == 200, r.text
    tables = {t["name"]: t for t in r.json()["tables"]}
    assert "studies" in tables and "accounts" in tables
    cols = {c["name"] for c in tables["studies"]["columns"]}
    assert {"study_uid", "modality", "status"} <= cols
    assert tables["accounts"]["rows"] >= 1  # admin 계정 존재


def test_db_tool_open(client, auth_headers, db, monkeypatch):
    from app.services.settings_service import get_setting, set_setting

    # 설정 없음 → 400 (이 테스트 이전에 키가 없어야 하므로 먼저 검사)
    if not (get_setting(db, "server.dbtool", default={}) or {}).get("path"):
        assert client.post("/api/insights/db-tool-open", headers=auth_headers).status_code == 400

    # 존재하지 않는 경로 → 400
    set_setting(db, "server.dbtool", {"path": "C:/no/such/dbtool.exe"}, scope="global")
    assert client.post("/api/insights/db-tool-open", headers=auth_headers).status_code == 400

    # 실재 파일 → Popen 분리 실행(모킹) + 감사 로그
    set_setting(db, "server.dbtool", {"path": sys.executable}, scope="global")
    calls: dict = {}

    def fake_popen(args, **kw):
        calls["args"] = args
        calls["kw"] = kw

    monkeypatch.setattr(insights.subprocess, "Popen", fake_popen)
    r = client.post("/api/insights/db-tool-open", headers=auth_headers)
    assert r.status_code == 200, r.text
    assert r.json()["ok"] is True
    assert calls["args"] == [sys.executable], "설정 경로 단일 인자만 — 임의 인자 주입 금지"
    logs = client.get("/api/insights/logs", headers=auth_headers,
                      params={"q": "db_tool_open"}).json()["items"]
    assert any(it["action"] == "db_tool_open" for it in logs)


# ────────────────────────────── 관리자 계정 등록 가드 ──────────────────────────────
def test_admin_account_creation_guard(client, auth_headers):
    # ① 시스템 관리자는 role=admin 계정을 등록할 수 있다
    uname = f"adm2_{uuid.uuid4().hex[:6]}"
    r = client.post("/api/admin/accounts", headers=auth_headers, json={
        "username": uname, "password": "adminpass12", "role": "admin",
    })
    assert r.status_code == 200, r.text
    aid = r.json()["id"]

    # ② 비관리자는 admin 계정을 등록할 수 없다(권한 게이트에서 차단)
    tname = f"tech2_{uuid.uuid4().hex[:6]}"
    client.post("/api/admin/accounts", headers=auth_headers, json={
        "username": tname, "password": "pass12345", "role": "technologist",
    })
    hdrs = _login(client, tname, "pass12345")
    r = client.post("/api/admin/accounts", headers=hdrs, json={
        "username": f"evil_{uuid.uuid4().hex[:6]}", "password": "pass12345", "role": "admin",
    })
    assert r.status_code == 403

    # 정리 — 추가 admin 계정 삭제(다른 테스트의 '마지막 관리자' 가정 보존)
    assert client.delete(f"/api/admin/accounts/{aid}", headers=auth_headers).status_code == 200
